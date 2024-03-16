from flask import Flask, render_template, request, redirect, url_for
import yolov5
import cv2
from PIL import Image
import easyocr
import tempfile
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Load YOLOv5 model
def load_model():
    return yolov5.load('licence-plate.pt')

# Set YOLOv5 model parameters
def set_model_parameters(model):
    model.conf = 0.25  # NMS confidence threshold
    model.iou = 0.45  # NMS IoU threshold
    model.agnostic = False  # NMS class-agnostic
    model.multi_label = False  # NMS multiple labels per box
    model.max_det = 1000  # maximum number of detections per image
    return model

# Perform inference
def inference(model, img):
    results = model(img)
    return results

# Parse results
def parse_results(results):
    predictions = results.pred[0]
    boxes = predictions[:, :4]  # x1, y1, x2, y2
    scores = predictions[:, 4]
    categories = predictions[:, 5]
    return boxes, scores, categories

# Find bounding box with highest score
def find_max_score_index(scores):
    return scores.argmax() or 0

# Extract bounding box with highest score
def extract_bounding_box(boxes, max_score_index):
    return map(int, boxes[max_score_index])

# Load original image
def load_original_image(img):
    original_image = cv2.imread(img)
    original_image = cv2.cvtColor(original_image, cv2.COLOR_BGR2RGB)
    return original_image

# Crop region of interest from original image
def crop_region_of_interest(original_image, x1, y1, x2, y2):
    return original_image[y1:y2, x1:x2]

# Convert cropped image to PIL format and save to temporary file
def convert_to_pil_and_save(cropped_image):
    cropped_image_pil = Image.fromarray(cropped_image)
    temp_img_path = tempfile.mktemp(suffix='.jpg')
    cropped_image_pil.save(temp_img_path)
    return temp_img_path

# Initialize EasyOCR reader
def initialize_easyocr():
    return easyocr.Reader(['en'])

# Perform OCR on cropped image
def perform_ocr(reader, temp_img_path):
    result = reader.readtext(temp_img_path)
    return result[0][1]

model = load_model()
model = set_model_parameters(model)
reader = initialize_easyocr()



# Initialize EasyOCR reader
reader = initialize_easyocr()

# Create an Excel workbook if it doesn't exist, otherwise load existing workbook
if not os.path.exists("ocr_results.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.append(["OCR Result"])  # Add header row if creating a new workbook
else:
    wb = load_workbook("ocr_results.xlsx")
    ws = wb.active

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            img_path = os.path.join('uploads', file.filename)
            file.save(img_path)
            results = inference(model, img_path)
            boxes, scores = parse_results(results)
            max_score_index = find_max_score_index(scores)
            x1, y1, x2, y2 = extract_bounding_box(boxes, max_score_index)
            original_image = load_original_image(img_path)
            cropped_image = crop_region_of_interest(original_image, x1, y1, x2, y2)
            temp_img_path = convert_to_pil_and_save(cropped_image)
            ocr_result = perform_ocr(reader, temp_img_path)
            ocr_clean = ocr_result.replace("*", " ").replace("@", " ").replace(".", " ").replace("'", " ").upper()
            ws.append([ocr_clean])
            wb.save("ocr_results.xlsx")
            return render_template('result.html', image_file=img_path, ocr_result=ocr_clean)
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
